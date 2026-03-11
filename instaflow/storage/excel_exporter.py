"""
Excel Exporter — Module 6.3

Exports extracted links to formatted .xlsx files using openpyxl.

Output:
  • Daily file:  exports/instaflow_export_YYYY-MM-DD.xlsx
  • Cumulative:  exports/instaflow_ALL_TIME.xlsx

Columns: #, Reel URL, Creator, DM Message, Original URL, Final URL,
         Redirect Hops, Method, Timestamp
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from instaflow.config.logging import get_logger
from instaflow.config.settings import get_settings

logger = get_logger(__name__)

COLUMNS = [
    "#",
    "Reel URL",
    "Creator",
    "DM Message",
    "Original URL",
    "Final URL",
    "Redirect Hops",
    "Method",
    "Timestamp",
]

# Header styling
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _ensure_workbook(file_path: Path) -> Workbook:
    """Load existing workbook or create a new one with headers."""
    if file_path.exists():
        wb = load_workbook(file_path)
        return wb

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Links"

    # Write headers with styling
    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(file_path)
    return wb


def _auto_width(ws: Any) -> None:
    """Auto-fit column widths based on content."""
    for col_idx in range(1, len(COLUMNS) + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            len(str(cell.value or ""))
            for cell in ws[letter]
        )
        ws.column_dimensions[letter].width = min(max_len + 4, 60)


def export_links(records: list[dict[str, Any]]) -> dict[str, str]:
    """
    Append extracted-link records to both the daily and cumulative Excel files.

    Parameters
    ----------
    records : list of dicts with keys matching COLUMNS (snake_case).

    Returns
    -------
    dict with 'daily_file' and 'cumulative_file' paths.
    """
    settings = get_settings()
    exports_dir = Path(settings.exports_dir)
    exports_dir.mkdir(parents=True, exist_ok=True)

    today = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d")
    daily_path = exports_dir / f"instaflow_export_{today}.xlsx"
    cumulative_path = exports_dir / "instaflow_ALL_TIME.xlsx"

    for file_path in (daily_path, cumulative_path):
        wb = _ensure_workbook(file_path)
        ws = wb["Extracted Links"]

        start_row = ws.max_row + 1

        for idx, record in enumerate(records, start=start_row - 1):
            row_num = ws.max_row + 1
            redirect_chain = record.get("redirect_chain") or []
            hop_count = len(redirect_chain)

            row_data = [
                idx,
                record.get("reel_url", ""),
                record.get("creator_username", ""),
                record.get("dm_message_text", ""),
                record.get("original_url", ""),
                record.get("final_url", ""),
                hop_count,
                record.get("extraction_method", ""),
                record.get("extracted_at", ""),
            ]

            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col_idx, value=str(value) if value else "")

        _auto_width(ws)
        wb.save(file_path)

    logger.info(
        "excel.exported",
        record_count=len(records),
        daily_file=str(daily_path),
    )

    return {
        "daily_file": str(daily_path),
        "cumulative_file": str(cumulative_path),
    }
